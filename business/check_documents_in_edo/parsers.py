import re
from pathlib import WindowsPath
from typing import List

from bs4 import BeautifulSoup, ResultSet
from bs4.element import Tag
from openpyxl import load_workbook, Workbook
from pdfminer.high_level import extract_text


class ExcelParser:

    def __init__(self, file_path: str = None, read_only: bool = True, write_only: bool = False):
        self.file_path = file_path
        self.read_only = read_only
        self.write_only = write_only
        self.__wb = None
        self.__ws = None
        self.__titles = None
        self.__titles_indexes = None

    def get_titles(self) -> list:
        """
        Получаем Заголовки листа.
        :return: Возвращаем массив заголовков
        """

        self.__init_excel_if_not_exist()

        return list(self.__ws.values)[0]

    def get_data(self) -> list:
        """
        Получить все данные
        :return: Возвращает все данные с листа.
        """

        self.__init_excel_if_not_exist()

        data = list()

        self.__titles = self.get_titles()
        self.__titles_indexes = self.__titles_indexing()

        for row in self.__ws.iter_rows(min_row=2, values_only=True):
            item = self.__data_indexing(row)
            data.append(item)

        return data

    def close_excel(self):
        """
        Закрывает ексель
        :return:
        """

        self.__wb.close()

    def __create_workbook(self):
        """
        Метод создает книгу
        Пока ни разу не использовался. Возможно пригодится вдальнейшем, но не факт
        :return:
        """

        self.__wb = Workbook(write_only=self.write_only)

    def __init_workbook(self):
        """
        Открывает эксель файл.
        :return:
        """

        self.__wb = load_workbook(filename=self.file_path, read_only=self.read_only)

    def __init_worksheet(self):
        """
        Выбираем активный Лист.
        :return:
        """

        self.__ws = self.__wb.active

    def __init_excel_if_not_exist(self):
        """
        Метод проверяет был ли инициализирован ексель файл, если нет то инициализирует
        :return:
        """

        if not self.__wb:
            self.__init_workbook()

        if not self.__ws:
            self.__init_worksheet()

    def __titles_indexing(self) -> dict:
        """
        Индексация заголовков
        :return: индексированный список заголовков
        """
        title_indexes = dict()

        for title in self.__titles:
            title_indexes[title] = self.__titles.index(title)

        return title_indexes

    def __data_indexing(self, row: list) -> dict:
        """
        Метод индексации строки из екселя
        :param row: строка в виде массива из екселя
        :return: Справочник вида "Заголовок: данные"
        """

        item = dict()

        for name in self.__titles:
            item[name] = row[self.__titles_indexes.get(name)]

        return item


class WebGuiParser:

    def parse_webgui_table(self, cdata: str, header_left_id: str, header_right_id: str, body_left_id: str, body_right_id: str) -> List[dict]:
        """
        Метод парсит webgui-евские таблицы, которые состоят из 4ех таблиц. P.S. Если ты читаешь это описание я думаю ты уже понял, что такое "таблицы, которые состоят из 4ех таблиц" .
        :param cdata: cdata ответа, из которого нужно парсить таблицу
        :param header_left_id: id заголовков левой части таблицы
        :param header_right_id: id заголовков правой части таблицы
        :param body_left_id: id левой стороны таблицы с данными
        :param body_right_id: id правой стороны таблицы с данными
        :return: список позиций в формате dict
        """

        html = self.__get_cdata_content(cdata)

        soup = BeautifulSoup(html, features='html.parser')

        # Очередной изврат одна таблица на карточке разбита на 4 разных в html...

        table_header_left = self.__get_html_table(soup=soup, attr='id', value=header_left_id)
        table_header_right = self.__get_html_table(soup=soup, attr='id', value=header_right_id)
        table_left = self.__get_html_table(soup=soup, attr='id', value=body_left_id)
        table_right = self.__get_html_table(soup=soup, attr='id', value=body_right_id)

        if not table_header_left or not table_header_right or not table_left or not table_right:
            raise AttributeError('%s. Не смог получить элементы по предоставленыи id' % self.parse_webgui_table.__qualname__)

        header_data_left = self.__parse_table_data(table_header_left, 'th')
        header_data_right = self.__parse_table_data(table_header_right, 'th')
        left_data = self.__parse_table_data(table_left, 'tr')
        right_data = self.__parse_table_data(table_right, 'tr')

        # Начинается вакханалия с объеденением этих таблиц в одну и сбор dict из всего этого непотребства
        # Регулярка, которая призвана вычищать к х.. все лишние пробелы
        pattern = re.compile(r'(\s){2,}')

        # Забираем все значения заголовков из td, убирая лишние пробелы
        headers = list(map(lambda x: re.sub(pattern, ' ', x[0].text).strip() if len(x) != 0 else 'Колонка с Чекбоксами', header_data_left + header_data_right))

        # Будущий нормальный список данных таблицы
        data = list()

        # Объединяем 2 таблицы в одну. Т.к. кол-во элементов одинаково, достаочно бежать циклом по right_data и забирать значения по индексу из left_data
        for i, item in enumerate(right_data):
            # Забираем все значения заголовков из td, убирая лишние пробелы
            data.append(list(map(lambda x: re.sub(pattern, ' ', x.text).strip(), left_data[i] + item)))

        return self.__create_dict_(headers, data)

    def _parse_details_listbox(self, cdata: str) -> List[dict]:
        """
        Парсит выпадающий список "Позиция" на карточке 45*-го над таблицей "Позиция подробно"
        :param cdata: текст ответа метода WebGui45ContractCard.init_card
        :return: список dict где dict формата : {data_item_key: '   <value>', name: '[ <Позиция> ] <Код материала>, <Краткий текст материала>'}
        """

        list_box_container_id = 'DYN_6000-LISTSAPLMEGUI_ei-scrl'

        html = self.__get_cdata_content(cdata)

        soup = BeautifulSoup(html, features='html.parser')

        container = soup.find(attrs={'id': list_box_container_id})

        list_box = container.find(attrs={'class': 'lsListbox__values'})

        list_box_values = list_box.find_all(attrs={'class': 'lsListbox__value'})

        return list(map(lambda x: dict(data_item_key=x.get('data-itemkey'), value=x.text), list_box_values))

    def get_value_by_attr(self, cdata: str, find_value: str, get_attr_name: str = 'value', find_attr: str = 'id') -> str:
        """
        Универсальный метод для парсинга значений по идентификатору поля
        :param cdata: текст ответа от запроса, из которого нужно забрать данные
        :param find_value: значения для атрибутта поиска
        :param find_attr: аттрибут по которому будет осуществляться поиск. По умолачию id
        :param get_attr_name: аттрибут по которому будет забираться значение в найденом элементе.По умолачию value
        :return: значение поля
        """

        html = self.__get_cdata_content(cdata)

        soup = BeautifulSoup(html, features='html.parser')

        tag = soup.find(attrs={find_attr: find_value})

        if not tag:
            raise AttributeError(f'%s. Не смог получить значение по атрибуту %s = %s' % (self.get_value_by_attr.__qualname__, find_attr, find_value))

        return tag.attrs.get(get_attr_name)

    def __parse_table_data(self, table: Tag, row_type: str) -> List[ResultSet]:
        """
        Метод парсит данные из таблицы HTML
        :param row_type: 'tr' or 'th'
        :param table: таблица для парсинга
        :return: масссив данных из таблицы
        """

        result = list()
        table_body = table.find('tbody')

        rows = table_body.find_all(row_type)

        for row in rows:
            result.append(row.find_all('td'))

        return result

    def __get_html_table(self, soup: BeautifulSoup, attr: str, value: str) -> Tag:
        """
        Ищет таблицу по значению атрибута
        :param attr: атрибут id, class, и тд.
        :param soup: экземпляр bs4.
        :param value: значение атрибути
        :return: Таблицу
        """

        return soup.find('table', attrs={attr: value})

    def __create_dict_(self, key_list: list, value_list: list):
        """
        Метод конвертит массивы с заголовками и данными в dict формата {"Ключ": "Значение"}
        :param key_list: Список заголовков\Ключей
        :param value_list: Список значений
        :return: dict
        """

        title_indexes = dict()
        result = list()

        # Порлучаем индексы заголовков формата {"Заголовок": 0(Индекс заголовка в массиве)}
        for key in key_list:
            title_indexes[key] = key_list.index(key)

        # Цикл по всем данным
        for value in value_list:

            item = dict()

            # Цикл по массиву(строке таблицы), в котором собирается dict вида {"Заголовок": "значение"}
            for name in key_list:
                item[name] = value[title_indexes.get(name)]

            result.append(item)

        return result

    def __get_cdata_content(self, cdata: str) -> str:
        """
        Парсит html из <![CDATA[ <html/> ]]>
        :param cdata: html обернутый в cdata
        :return: html
        """

        # Костылина ибо parse_between возвращает NONE слишком большой ответ для него, а использовать сразу суп не получается т.к. приходит говно, а не html
        html = re.sub('^(.+?)<content-update id="webguiPage0"><!\[CDATA\[', '', cdata)

        if not html:
            raise ValueError('Лезь в WebGuiParser._parse_order_card_table.__get_cdata_content не смог обрезать первую часть ответа')

        html = re.sub(']]></content-update(.+?)$', '', html)

        if not html:
            raise ValueError('Лезь в WebGuiParser._parse_order_card_table.__get_cdata_content не смог обрезать вторую часть ответа')

        return html


class PDFParser:

    def __init__(self, pdf_path: WindowsPath = None):
        self.path = pdf_path

    # noinspection PyTypeChecker
    def get_text(self) -> str:
        """
        Метод получает весь текст с пдф файла
        :return:
        """

        if not self.path:
            raise ValueError('%s. Укажите pdf_path при инициализации класса' % self.get_text.__qualname__)

        try:
            with open(self.path, 'rb') as f:
                text = extract_text(f)
            return text
        except Exception as ex:
            raise Exception('%s. Не удалось получить данные из Файла. %s' % (self.get_text.__qualname__, ex))
